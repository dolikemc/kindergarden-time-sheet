from datetime import datetime, date, timedelta
from typing import Dict, Iterator, Union, List
from calendar import Calendar
import holidays
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, Color, Side
from openpyxl.styles.builtins import styles
from openpyxl.cell.cell import Cell
from yaml import load, SafeLoader
from pathlib import Path

date_types = ['normal', 'weekend', 'holiday', 'special']


class DataRow:
    date: date
    type: int  # key of data_types
    name: str


class Configurator:
    def __init__(self):
        with (Path.cwd() / 'config.yaml').open() as config_file:
            self.config: Dict = load(config_file.buffer, SafeLoader)


class DateHandler:
    _worksheet: Worksheet
    _config: Dict
    _calendar: Calendar
    _holidays: holidays
    _special_days: Dict

    def __init__(self, sheet: Union[Worksheet, None], config: Configurator) -> None:
        if sheet:
            self._worksheet = sheet
        self._config = config.config
        self.font = Font(size=16)
        self.side = Side(style='medium', color='00000000')

        self._calendar = Calendar(firstweekday=0)
        self._holidays: Dict = holidays.country_holidays(
            country=self._config.get('country', 'DE'),
            subdiv=self._config.get('subdiv', 'BY'),
            years=self._config.get('year', datetime.today().year)
        )
        self._special_days = {}
        for name in self._config.get('holidays', [{'name': ''}]):
            for date_record in name.get('dates', []):
                month_day = datetime.strptime(date_record, self._config.get('format', '%d/%m'))
                key_date = date(year=self._config.get('year', datetime.today().year),
                                month=month_day.month, day=month_day.day)
                self._special_days[key_date.isoformat()] = name.get('name', '')

    def year_iterator(self) -> Iterator[DataRow]:
        day = date(year=self._config.get('year', date.today().year), month=1, day=1)
        return_row = DataRow()
        for i in range(0, (date(year=self._config.get('year', date.today().year) + 1, month=1, day=1) -
                           date(year=self._config.get('year', date.today().year), month=1, day=1)).days):
            if day in self._holidays:
                return_row.date, return_row.type, return_row.name = day, 2, self._config.get(
                    'holiday', 'Holiday')
            elif day.weekday() > 4:
                return_row.date, return_row.type, return_row.name = day, 1, ''
            elif day.isoformat() in self._special_days.keys():
                return_row.date, return_row.type, return_row.name = day, 3, self._special_days[
                    day.isoformat()]
            else:
                return_row.date, return_row.type = day, 0
            yield return_row
            day += timedelta(days=1)

    def add_row(self, hours: List, stop: str = '') -> int:
        if len(hours) < 5:
            raise Exception('you have to provide hours for each weekday')

        dv = self.create_header()
        month_day = self.determine_stop_date(stop)
        for index, day_row in enumerate(self.year_iterator()):

            # stop criteria for a member
            if day_row.date > month_day:
                return 1

            if day_row.type:
                for c in range(1, 7):
                    self._worksheet.cell(row=index + 2, column=c).style = \
                        self._config.get('styles', ['Output', 'Output', 'Output'])[day_row.type - 1]
                self._worksheet.cell(row=index + 2, column=6, value=day_row.name)
                # self._worksheet.cell(row=index + 2, column=5, value="00:00")

            else:
                self._worksheet.cell(row=index + 2, column=4, value='00:00')
                self._worksheet.cell(row=index + 2, column=4).number_format = '[hh]:mm'

                self._worksheet.cell(row=index + 2, column=3, value=f"0{hours[day_row.date.weekday()]}:00")
                self._worksheet.cell(
                    row=index + 2, column=5,
                    value=f'=IF(AND(A{index + 2}<TODAY()-2,C{index + 2}<>"",F{index + 2}=""),'
                          f'IF(C{index + 2}=0,D{index + 2}*1.5,'
                    #      f'D{index + 2}-C{index + 2}'
                          f'TEXT(ABS(D{index + 2}-C{index + 2}),'
                          f'IF(_xlfn.NUMBERVALUE(D{index + 2})<_xlfn.NUMBERVALUE(C{index + 2}),"-","")&"[hh]:mm")'
                          f'),"")')
                self._worksheet.cell(
                    row=index + 2, column=17,
                    value=f'=IF(AND(A{index + 2}<TODAY()-2,C{index + 2}<>"",F{index + 2}=""),'
                          f'IF(C{index + 2}=0,D{index + 2}*1.5,'
                          f'D{index + 2}-C{index + 2}'
                          f'),"")')

                dv.add(self._worksheet.cell(row=index + 2, column=6))

            self._worksheet.cell(row=index + 2, column=3).number_format = '[hh]:mm'
            self._worksheet.cell(row=index + 2, column=1, value=day_row.date)
            self._worksheet.cell(row=index + 2, column=1).number_format = 'd/m'
            self._worksheet.cell(row=index + 2, column=2, value=day_row.date.strftime('%a'))
            self._worksheet.row_dimensions[index + 2].height = 25

            for c in range(1, 7):
                self._worksheet.cell(row=index + 2, column=c).font = self.font
                self._worksheet.cell(row=index + 2, column=c).alignment = Alignment(horizontal="center")

            # if not self._worksheet.cell(row=index + 2, column=4).value:
            #    self._worksheet.cell(row=index + 2, column=5).font = Font(
            #        color=self._worksheet.cell(row=index + 2, column=5).fill.fgColor,
            #        # '00FFFFCC',
            #        size=16)

        return self._worksheet.max_row

    def determine_stop_date(self, stop):
        if stop:
            month_day = datetime.strptime(stop, self._config.get('format', '%d/%m'))
        else:
            month_day = date(year=self._config.get('year', datetime.today().year), month=12, day=31)
        return date(year=self._config.get('year', datetime.today().year),
                    month=month_day.month,
                    day=month_day.day)

    def create_header(self):
        dv = DataValidation(type='list', formula1='"krank,urlaub,kindkrank,fortbildung"')
        self._worksheet.add_data_validation(dv)
        header = [('A1', 'Datum'), ('B1', 'Wochentag'), ('C1', 'Soll'), ('D1', 'Ist'), ('E1', 'Saldo'),
                  ('F1', 'Abwesenheit')]
        for cell in header:
            self._worksheet[cell[0]] = cell[1]
            self._worksheet[cell[0]].style = styles['Title']
            self._worksheet.column_dimensions[cell[0][0:1]].width = 20

        return dv

    def summary(self, name: str):

        for i in ('G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P'):
            self._worksheet.column_dimensions[i].width = 20
        # header for name
        self._worksheet.merge_cells(start_row=8, start_column=8, end_row=8, end_column=16)
        self.set_cell_std_format(from_row=8, from_column=8, style_number=4, text=f'Zusammenfassung {name}')

        # header holiday
        self._worksheet.merge_cells(start_row=10, start_column=8, end_row=10, end_column=12)
        self.set_cell_std_format(from_row=10, from_column=8, style_number=3, text='Urlaubstage')
        self.set_cell_std_format(from_row=11, from_column=8, style_number=3,
                                 text=f"Rest {self._config.get('year', datetime.today().year) - 1}")
        self.set_cell_std_format(from_row=11, from_column=9, style_number=3,
                                 text=f"Soll {self._config.get('year', datetime.today().year)}")
        self.set_cell_std_format(from_row=11, from_column=10, style_number=3, text="Summe")
        self.set_cell_std_format(from_row=11, from_column=11, style_number=3, text="Genommen")
        self.set_cell_std_format(from_row=11, from_column=12, style_number=3, text="Offen")

        # header overtime
        self._worksheet.merge_cells(start_row=10, start_column=14, end_row=10, end_column=16)
        self.set_cell_std_format(from_row=10, from_column=14, style_number=3, text='Überstunden')
        self.set_cell_std_format(from_row=11, from_column=14, style_number=3,
                                 text=f"Rest {self._config.get('year', datetime.today().year) - 1}")
        self.set_cell_std_format(from_row=11, from_column=15, style_number=3,
                                 text=f"{self._config.get('year', datetime.today().year)}")
        self.set_cell_std_format(from_row=11, from_column=16, style_number=3, text="Summe")

        # holiday cells
        self.set_cell_std_format(from_row=12, from_column=8)
        self.set_cell_std_format(from_row=12, from_column=9, text='30')
        self.set_cell_std_format(from_row=12, from_column=10, text="=H12+I12")
        self.set_cell_std_format(from_row=12, from_column=11, text='=COUNTIF(F2:F367,"urlaub")')
        self.set_cell_std_format(from_row=12, from_column=12, text="=J12-K12")

        # overtiem cells
        # =TEXT(ABS(SUMME((-1+2*(LINKS(G11:G22)<>"-"))*(RECHTS(0&G11:G22;5))));WENN(SUMME((-1+2*(LINKS(G11:G22)<>"-"))*(RECHTS(0&G11:G22;5)))<0;"-";)&"[hh]:mm:ss")
        self.set_cell_std_format(from_row=12, from_column=14, text="00:00", number_format='[hh]":"mm')
        self.set_cell_std_format(from_row=12, from_column=15,
                                 text=
                                 # '=TEXT(ABS(SUM((-1+2*(LEFT(E2:E367)<>"-"))*(RIGHT(0&E2:E367,5)))),'
                                 #     'IF(SUM((-1+2*(LEFT(E2:E367)<>"-"))*(RIGHT(0&E2:E367,5)))<0,"-",)&"[hh]:mm")',
                                 '=TEXT(ABS(SUM(Q2:Q367)),IF(_xlfn.NUMBERVALUE(SUM(Q2:Q367))<0,"-","")&"[hh]:mm")',
                                 number_format='[hh]":"mm')
        self.set_cell_std_format(from_row=12, from_column=16,
                                 # =IF(LEFT(N12)="-",-xlfn.NUMBERVALUE(RIGHT(0&N12,5)),xlfn.NUMBERVALUE(N12))
                                 text='=TEXT(ABS(SUM(Q2:Q367)+Q1),'
                                      'IF(_xlfn.NUMBERVALUE(Q1+SUM(Q2:Q367))<0,"-","")&"[hh]:mm")',
                                 # "=N12+O12",
                                 number_format='[hh]":"mm')

        self.set_cell_std_format(from_row=1, from_column=17,
                                 text='=IF(LEFT(N12)="-",_xlfn.NUMBERVALUE(RIGHT(0&N12,5))*-1,_xlfn.NUMBERVALUE(N12))')

        # header training
        self._worksheet.merge_cells(start_row=14, start_column=8, end_row=14, end_column=10)
        self.set_cell_std_format(from_row=14, from_column=8, style_number=3, text='Fortbildung')
        self.set_cell_std_format(from_row=15, from_column=8, style_number=3,
                                 text=f"Soll {self._config.get('year', datetime.today().year)}")
        self.set_cell_std_format(from_row=15, from_column=9, style_number=3, text=f"Genommen")
        self.set_cell_std_format(from_row=15, from_column=10, style_number=3, text="Offen")

        # header sick
        self.set_cell_std_format(from_row=14, from_column=12, style_number=3, text="Krankheitstage")
        self.set_cell_std_format(from_row=15, from_column=12, style_number=3,
                                 text=f"{self._config.get('year', datetime.today().year)}")

        # header child sick
        self._worksheet.merge_cells(start_row=14, start_column=14, end_row=14, end_column=16)
        self.set_cell_std_format(from_row=14, from_column=14, style_number=3, text='Kinder Krankeitstage')
        self.set_cell_std_format(from_row=15, from_column=14, style_number=3,
                                 text=f"Soll {self._config.get('year', datetime.today().year)}")
        self.set_cell_std_format(from_row=15, from_column=15, style_number=3, text="Genommen")
        self.set_cell_std_format(from_row=15, from_column=16, style_number=3, text="Offen")

        # training cells
        self.set_cell_std_format(from_row=16, from_column=8, text='5')
        self.set_cell_std_format(from_row=16, from_column=9, text='=COUNTIF(F2:F367,"fortbildung")')
        self.set_cell_std_format(from_row=16, from_column=10, text="=H16-I16")

        # sick cells
        self.set_cell_std_format(from_row=16, from_column=12, text='=COUNTIF(F2:F367,"krank")')

        # child sick cells
        self.set_cell_std_format(from_row=16, from_column=14, text='10')
        self.set_cell_std_format(from_row=16, from_column=15, text='=COUNTIF(F2:F367,"kindkrank")')
        self.set_cell_std_format(from_row=16, from_column=16, text="=N16-O16")

        self._worksheet.column_dimensions['Q'].hidden = True

    def set_cell_std_format(self, from_row: int, from_column: int, style_number: int = None, text: str = None,
                            number_format: str = None) -> Cell:
        cell: Cell = self._worksheet.cell(row=from_row, column=from_column)
        if style_number:
            cell.style = self._config.get('styles', ['Output', 'Output', 'Output', 'Output'])[style_number]
        if text:
            cell.value = text
        if number_format in ['d/m', '0.00', '[hh]":"mm']:
            cell.number_format = number_format
        cell.font = self.font
        cell.alignment = Alignment(horizontal="center")
        return cell
