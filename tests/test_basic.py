import unittest

from openpyxl import Workbook, load_workbook
from src.DataRow import DateHandler, Configurator


class BasicTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.config = Configurator()

    def test_init(self):
        self.assertIsInstance(DateHandler(None, config=self.config), DateHandler)

    def test_generate(self):
        dr = DateHandler(None, config=self.config)
        self.assertGreaterEqual(len([x for x in dr.year_iterator()]), 356)

    def test_add_row(self):
        wb = Workbook()
        dr = DateHandler(wb.active, config=self.config)
        self.assertTrue(dr)
        self.assertTrue(dr.add_row(hours=[5, 5, 7, 9, 2]))
        wb.save('test.xlsx')

    def test_validator(self):
        wb = load_workbook('test2.xlsx')
        print(wb.active.cell(2, 1).number_format)
        print(wb.active['E11'].value)
        print(wb.active['I10'].value)
        print(wb.active['O12'].value)
        print(wb.active.column_dimensions['Q'])
        print(wb.active['Q1'])

        for datat in wb.active.data_validations.dataValidation:
            print(datat)
